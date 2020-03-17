import random
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from datetime import datetime




wb = Workbook()

dest_filename = 'FMC1_AAT.xlsx'

#ws1 = wb.active
#ws1.title = "range names"

#for row in range(1, 8):

#    ws1.append(range(600))


ws3 = wb.create_sheet(title="Data")

#val = ['G14','G15','G16','G17', 'G18','G19', 'G20', 'G21', 'H14','H15','H16','H17', 'H18','H19', 'H20', 'H21','I14','I15','I17','I18','I19','I21']
#for i, item in enumerate(val):
#    if item < 10:
#        val[i] = 99
#print(val)
#for i in range(100):
timestamp = 1570736760
#    for row in range(14, 21):
def casnatext(timestamp):
    date_time =datetime.fromtimestamp(timestamp)
    return date_time.strftime("%H:%M:%S")
def datumnatext(timestamp):
    date_time =datetime.fromtimestamp(timestamp)
    return date_time.strftime("%d.%m.%Y")




#        for col in range(8,9):
for i in range(1139):
#    timestamp = 1570736760
#    for timestamp in range(1570736760,1573422960,1570740960):

#        date_time = datetime.fromtimestamp(timestamp)
#       d = date_time.strftime("%X")
#        print("Output 3:", d)
    cislo = [4.106, 4.107]
    vyb_cislo = random.choice(cislo)

    cislo2 = [2048, 2049]
    vyb_cislo2 = random.choice(cislo2)

    cislo3 = [9.056, 9.066, 9.067, 9.068, 9.071, 9.072]
    vyb_cislo3 = random.choice(cislo3)

    ws3.cell(row=i * 8 + 1694, column=1, value=("{}.CYKLUS".format(i+202)))
    ws3.cell(row=i * 8 + 1695, column=1, value=("{}.CYKLUS".format(i+202)))
    ws3.cell(row=i * 8 + 1696, column=1, value=("{}.CYKLUS".format(i+202)))
    ws3.cell(row=i * 8 + 1697, column=1, value=("{}.CYKLUS".format(i+202)))
    ws3.cell(row=i * 8 + 1698, column=1, value=("{}.CYKLUS".format(i+202)))
    ws3.cell(row=i * 8 + 1699, column=1, value=("{}.CYKLUS".format(i+202)))
    ws3.cell(row=i * 8 + 1700, column=1, value=("{}.CYKLUS".format(i+202)))
    ws3.cell(row=i * 8 + 1701, column=1, value=("{}.CYKLUS".format(i+202)))

    ws3.cell(row=i * 8 + 1694, column=2, value=i+202)
    ws3.cell(row=i * 8 + 1695, column=2, value=i+202)
    ws3.cell(row=i * 8 + 1696, column=2, value=i+202)
    ws3.cell(row=i * 8 + 1697, column=2, value=i+202)
    ws3.cell(row=i * 8 + 1698, column=2, value=i+202)
    ws3.cell(row=i * 8 + 1699, column=2, value=i+202)
    ws3.cell(row=i * 8 + 1700, column=2, value=i+202)
    ws3.cell(row=i * 8 + 1701, column=2, value=i+202)

    ws3.cell(row=i * 8 + 1694, column=3, value='DATE')
    ws3.cell(row=i * 8 + 1695, column=3, value='DATE')
    ws3.cell(row=i * 8 + 1696, column=3, value='DATE')
    ws3.cell(row=i * 8 + 1697, column=3, value='DATE')
    ws3.cell(row=i * 8 + 1698, column=3, value='DATE')
    ws3.cell(row=i * 8 + 1699, column=3, value='DATE')
    ws3.cell(row=i * 8 + 1700, column=3, value='DATE')
    ws3.cell(row=i * 8 + 1701, column=3, value='DATE')

    ws3.cell(row=i * 8 + 1694, column=4, value=(datumnatext(timestamp+i*4260)))
    ws3.cell(row=i * 8 + 1695, column=4, value=(datumnatext(timestamp+i*4260)))
    ws3.cell(row=i * 8 + 1696, column=4, value=(datumnatext(timestamp+i*4260)))
    ws3.cell(row=i * 8 + 1697, column=4, value=(datumnatext(timestamp+i*4260)))
    ws3.cell(row=i * 8 + 1698, column=4, value=(datumnatext(timestamp+120+i*4260)))
    ws3.cell(row=i * 8 + 1699, column=4, value=(datumnatext(timestamp+120+i*4260)))
    ws3.cell(row=i * 8 + 1700, column=4, value=(datumnatext(timestamp+4200+i*4260)))
    ws3.cell(row=i * 8 + 1701, column=4, value=(datumnatext(timestamp+4200+i*4260)))

    ws3.cell(row=i * 8 + 1694, column=5, value='TIME')
    ws3.cell(row=i * 8 + 1695, column=5, value='TIME')
    ws3.cell(row=i * 8 + 1696, column=5, value='TIME')
    ws3.cell(row=i * 8 + 1697, column=5, value='TIME')
    ws3.cell(row=i * 8 + 1698, column=5, value='TIME')
    ws3.cell(row=i * 8 + 1699, column=5, value='TIME')
    ws3.cell(row=i * 8 + 1700, column=5, value='TIME')
    ws3.cell(row=i * 8 + 1701, column=5, value='TIME')

    ws3.cell(row=i * 8 + 1694, column=6, value=(casnatext(timestamp+i*4260)))
    ws3.cell(row=i * 8 + 1695, column=6, value=(casnatext(timestamp+i*4260)))
    ws3.cell(row=i * 8 + 1696, column=6, value=(casnatext(timestamp+i*4260)))
    ws3.cell(row=i * 8 + 1697, column=6, value=(casnatext(timestamp+i*4260)))
    ws3.cell(row=i * 8 + 1698, column=6, value=(casnatext(timestamp+120+i*4260)))
    ws3.cell(row=i * 8 + 1699, column=6, value=(casnatext(timestamp+120+i*4260)))
    ws3.cell(row=i * 8 + 1700, column=6, value=(casnatext(timestamp+4200+i*4260)))
    ws3.cell(row=i * 8 + 1701, column=6, value=(casnatext(timestamp+4200+i*4260)))

    ws3.cell(row=i * 8 + 1694,column=7, value='Ubus')
    ws3.cell(row=i * 8 + 1695, column=7, value='Usup')
    ws3.cell(row=i * 8 + 1696, column=7, value='TEMPERATURE')
    ws3.cell(row=i * 8 + 1697, column=7, value='Setpoint')
    ws3.cell(row=i * 8 + 1698, column=7, value='Status')
    ws3.cell(row=i * 8 + 1699, column=7, value='FuelFlow')
    ws3.cell(row=i * 8 + 1700, column=7, value='PASS')
    ws3.cell(row=i * 8 + 1701, column=7, value='Status')

    ws3.cell(row=i * 8 + 1694, column=8, value=241)
    ws3.cell(row=i * 8 + 1695, column=8, value=random.uniform(24.0, 24.2))
    ws3.cell(row=i * 8 + 1696, column=8, value=random.uniform(118.0, 124.0))
    ws3.cell(row=i * 8 + 1697, column=8, value=129.9989013672)
    ws3.cell(row=i * 8 + 1698, column=8, value=20032)
    ws3.cell(row=i * 8 + 1699, column=8, value=random.uniform(74.70, 75.40))
    ws3.cell(row=i * 8 + 1701, column=8, value=24128)

    ws3.cell(row=i * 8 + 1694, column=9, value='OK')
    ws3.cell(row=i * 8 + 1695, column=9, value='OK')
    ws3.cell(row=i * 8 + 1697, column=9, value='FlowFB')
    ws3.cell(row=i * 8 + 1698, column=9, value='Failure1')
    ws3.cell(row=i * 8 + 1699, column=9, value='FlowFb')
    ws3.cell(row=i * 8 + 1701, column=9, value='Failure1')

    ws3.cell(row=i * 8 + 1697, column=10, value=vyb_cislo)
    ws3.cell(row=i * 8 + 1698, column=10, value=2048)
    ws3.cell(row=i * 8 + 1699, column=10, value=random.uniform(13.620, 13.733))
    ws3.cell(row=i * 8 + 1701, column=10, value=vyb_cislo2)

    ws3.cell(row=i * 8 + 1697, column=11, value='FtempFb')
    ws3.cell(row=i * 8 + 1698, column=11, value='Failure2')
    ws3.cell(row=i * 8 + 1699, column=11, value='OK')
    ws3.cell(row=i * 8 + 1701, column=11, value='Failure2')

    ws3.cell(row=i * 8 + 1697, column=12, value=vyb_cislo3)
    ws3.cell(row=i * 8 + 1698, column=12, value=0)
    ws3.cell(row=i * 8 + 1701, column=12, value=0)

    ws3.cell(row=i * 8 + 1697, column=13, value='OK')
    ws3.cell(row=i * 8 + 1698, column=13, value='OK')
    ws3.cell(row=i * 8 + 1701, column=13, value='OK')

#            ws3['G14'] = 'Ubus'
#            ws3['G15'] = 'Usup'
#            ws3['G16'] = 'TEMPERATURE'
#            ws3['G17'] = 'Setpoint'
#            ws3['G18'] = 'Status'
#            ws3['G19'] = 'FuelFlow'
#            ws3['G20'] = 'PASS'
#            ws3['G21'] = 'Status'
#            ws3['H14'] = 241
#            ws3['H15'] = random.uniform(24.0,24.2)
#            ws3['H16'] = random.uniform(118.0, 124.0)
#            ws3['H17'] = 129.9989013672
#            ws3['H18'] = 24128
#            ws3['H19'] = random.uniform(74.70, 75.40)
#            ws3['H21'] = 24128
#            ws3['I14'] = 'OK'
#            ws3['I15'] = 'OK'

#            ws3['I17'] = 'FlowFB'
#            ws3['I18'] = 'Failure1'
#            ws3['I19'] = 'FlowFb'

 #           ws3['I21'] = 'Failure1'
 #           _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
#print(ws3['AA10'].value)

wb.save(filename=dest_filename)

#from openpyxl import load_workbook

#wb = load_workbook(filename = 'EXCELfile.xlsx')
#sheet_ranges = wb['Sheet2']
#print(sheet_ranges['A18'].value)

#for row in range

#import os
#print(os.getcwd())
#print("Current working dir : %s" % os.getcwd())

#file = open('pokus.txt', 'r', encoding='utf-8')
#contents = file.read()
#file.close()
#print(contents)

#x=('Hello world')
#print(x)
#print("This line will be printed.")

#from openpyxl import Workbook
#from openpyxl.utils import get_column_letter

#wb = Workbook()

#dest_filename = 'empty_book.xlsx'

#ws1 = wb.active
#ws1.title = "range names"

#for row in range(1, 8):

#   ws1.append(range(600))

#ws2 = wb.create_sheet(title="Pi")

#ws2['F5'] = 3.14

#ws3 = wb.create_sheet(title="Data")
#for row in range(10, 20):

#    for col in range(27, 54):

#            _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
#print(ws3['AA10'].value)

#wb.save(filename=dest_filename)